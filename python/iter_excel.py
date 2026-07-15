#!/usr/bin/env python3
# iter_excel.py
# Embedded python works in a shared memory with IRIS. 
# Thus direct calls to IRIS classmethods, globals and tables are available via iris lib.
import argparse
import iris
import openpyxl
import os
import sys

def iterate_excel_all_sheets(file_path):
    """
    Iterates over all sheets in an Excel file and prints non-empty cells.
    :param file_path: Path to the Excel file (.xlsx)
    """
    try:
        # Validate file existence
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Validate file extension
        if not file_path.lower().endswith(".xlsx"):
           raise ValueError("Only .xlsx files are supported.")

        # Initialize global reference
        gref = iris.gref('^wbook')

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        # Loop through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")

            # Iterate over all rows and cells
            for row in sheet.iter_rows():
                for cell in row:
                    # Skip empty cells (None or empty string after stripping)
                    if cell.value is None:
                        continue
                    if isinstance(cell.value, str) and cell.value.strip() == "":
                        continue
                    gvalue = str(cell.value)
                    print(f"Cell {cell.coordinate}: {cell.value}")
                    gref.set([sheet_name, cell.coordinate], gvalue)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        raise


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Iterate over cells in Excel workbook.")
    parser.add_argument("excel_file", type=str, help="xlsx Excel workbook file path")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose mode for additional output.")
    args = parser.parse_args()

    if args.verbose:
        print("Verbose mode enabled. Preparing to iterate...")

    excel_file = args.excel_file
    iterate_excel_all_sheets(excel_file)

    if args.verbose:
        print("Iterating completed successfully.")
